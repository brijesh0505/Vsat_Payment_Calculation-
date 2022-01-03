from openpyxl import Workbook, load_workbook
import datetime
from datetime import datetime

wb = load_workbook('test_2.xlsx')
sh1=wb['Sheet1']

i=2
j=3
k=sh1.max_column
l=sh1.max_row
while i<=l:
    v1=sh1.cell(i,j).value
    j+=1
    v2 = sh1.cell(i, j).value
    j+= 1
    v3 = sh1.cell(i, j).value
    j +=1
    v=v1+v2+v3
    sh1.cell(i,j).value=v
    j-=3
    i+=1

wb.save('test_2.xlsx')


