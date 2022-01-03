from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import datetime
from datetime import datetime
import openpyxl as px



wb = load_workbook('test_1.xlsx')
sh1=wb['Sheet1']
ws = wb.active

i=2
j=4
k=sh1.max_row
l=2
fill_pattern_flase=PatternFill(patternType='solid',fgColor='FFD700')

while i<=k:
    value_1=sh1.cell(i,j).value
    value_2=sh1.cell(i,j+1).value
    sh1.cell(i,j+3).value=(value_2-value_1)*24
    date_1=sh1.cell(i,l).value
    date_2=sh1.cell(i,l+1).value
    sh1.cell(i,j+4).value = date_1.strftime("%m")
    sh1.cell(i,j+5).value = date_2.strftime("%m")

    # true or false for month value
    if sh1.cell(i,j+4).value==sh1.cell(i,j+5).value:
        sh1.cell(i,j+6).value='TRUE'
    else:
        sh1.cell(i,j+6).value='FALSE'

    #color change for false value
    if sh1.cell(i,j+6).value=='FALSE':
        sh1.cell(i,j+6).fill = fill_pattern_flase
    else:
        pass

    # for total hours calculation in month
    if sh1.cell(i,j+4).value=="01":
        sh1.cell(i, j + 7).value =31*24
    elif sh1.cell(i,j+4).value=="02":
        sh1.cell(i, j + 7).value =28*24
    elif sh1.cell(i,j+4).value=="03":
        sh1.cell(i, j + 7).value =31*24
    elif sh1.cell(i,j+4).value=="04":
        sh1.cell(i, j + 7).value =30*24
    elif sh1.cell(i,j+4).value=="05":
        sh1.cell(i, j + 7).value =31*24
    elif sh1.cell(i,j+4).value=="06":
        sh1.cell(i, j + 7).value =30*24
    elif sh1.cell(i,j+4).value=="07":
        sh1.cell(i, j + 7).value =31*24
    elif sh1.cell(i,j+4).value=="08":
        sh1.cell(i, j + 7).value =31*24
    elif sh1.cell(i,j+4).value=="09":
        sh1.cell(i, j + 7).value =30*24
    elif sh1.cell(i,j+4).value=="10":
        sh1.cell(i, j + 7).value =31*24
    elif sh1.cell(i,j+4).value=="11":
        sh1.cell(i, j + 7).value =30*24
    elif sh1.cell(i,j+4).value=="12":
        sh1.cell(i, j + 7).value =31*24
    else:
        pass

    i+=1

# for adding filters
ws.auto_filter.ref = ws.dimensions

# save file
wb.save('test_1.xlsx')


# convert file xlsx to csv
import pandas as pd
read_file=pd.read_excel(r'test_1.xlsx')
read_file.to_csv(r'test_1.csv',index=None,header=True)


# data filtering using csv
df=pd.read_csv("test_1.csv")
a=df.pivot(index=["vsat_id", "sap_code"], columns=["month_diff_1"],values="diff_in_hours").fillna(0)
file_name='test_2.xlsx'
a.to_excel(file_name)




